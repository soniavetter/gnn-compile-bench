# -*- coding: utf-8 -*-
"""
Post-processing script that reads all result folders into the multi-sheet XLSX workbook.

Usage:
   python analyze_results_v12.py \
    -a results_auto \
    -t results_true \
    -f results_false \
    -n results_no_fixes \
    -g results_gin_300epochs \
    -o gnn_benchmark_results.xlsx
"""

import json
import statistics
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Color palette
GREEN_FILL   = PatternFill("solid", start_color="C6EFCE")
RED_FILL     = PatternFill("solid", start_color="FFC7CE")
ORANGE_FILL  = PatternFill("solid", start_color="FFEB9C")
GREY_FILL    = PatternFill("solid", start_color="D9D9D9")
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
SUBHDR_FILL  = PatternFill("solid", start_color="2E75B6")
SUBHDR2_FILL = PatternFill("solid", start_color="4472C4")
ALT_ROW_FILL = PatternFill("solid", start_color="EDF2F8")
SUMM_FILL    = PatternFill("solid", start_color="FFF2CC")
SUMM2_FILL   = PatternFill("solid", start_color="E2EFDA")
NO_FILL      = PatternFill()

HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
BODY_FONT    = Font(name="Arial", size=9)
BOLD_FONT    = Font(name="Arial", bold=True, size=9)
ERROR_FONT   = Font(name="Arial", bold=True, color="9C0006", size=9)
SUMM_FONT    = Font(name="Arial", bold=True, size=9, italic=True)

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# Compile modes
COMPILE_MODES = ["eager", "default", "reduce-overhead", "max-autotune", "max-autotune-no-cudagraphs"]
MODE_SHORT = {
    "eager":                      "eager",
    "default":                    "default",
    "reduce-overhead":            "reduce-OH",
    "max-autotune":               "max-AT",
    "max-autotune-no-cudagraphs": "max-AT-noCG",
}

# Canonical experiment ordering (matches paper table)
FW_ORDER    = ["pyg", "dgl"]
MODEL_ORDER = ["gcn", "graphsage", "gat", "gin", "rgcn", "distmult"]

def _fw_rank(fw):   return FW_ORDER.index(fw)      if fw    in FW_ORDER    else len(FW_ORDER)
def _model_rank(m): return MODEL_ORDER.index(m)    if m     in MODEL_ORDER  else len(MODEL_ORDER)
def exp_sort_key(k):  return (_fw_rank(k[0]), _model_rank(k[1]), k[2])
def pair_sort_key(p): return (_model_rank(p[0]), p[1])

# Metrics 
# (display_name, json_key, unit, lower_is_better)   lower_is_better=None -> neutral
METRICS = [
    ("Infer Latency Median (ms)",      "inference_latency_median_ms",           "ms",  True),
    ("Infer Latency IQR (ms)",         "inference_latency_iqr_ms",              "ms",  True),
    ("Infer Latency Std (ms)",         "inference_latency_std_ms",              "ms",  True),
    ("Infer Throughput (nodes/s)",     "throughput_inference_nodes_per_s",      "n/s", False),
    ("Train Throughput (nodes/s)",     "throughput_train_nodes_per_s",          "n/s", False),
    ("Train Throughput (edges/s)",     "throughput_train_edges_per_s",          "e/s", False),
    ("Peak GPU Infer (MB)",            "peak_gpu_memory_inference_mb",          "MB",  True),
    ("Peak GPU Train (MB)",            "peak_gpu_memory_train_mb",              "MB",  True),
    ("Median Epoch Time (s)",          "median_epoch_time_s",                   "s",   True),
    ("Mean Epoch Time (s)",            "mean_epoch_time_s",                     "s",   True),
    ("Compile Time (s)",               "compile_time_s",                        "s",   True),
    ("Total Compile Tax (s)",          "total_compile_tax_s",                   "s",   True),
    ("Speedup vs Eager",               "speedup_vs_eager",                      "x",   False),
    ("Train Speedup vs Eager",         "train_speedup_vs_eager",                "x",   False),
    ("Compile OH (equiv calls)",       "compile_overhead_equivalent_calls",     "",    True),
    ("Graph Capture Rate (%)",         "graph_capture_rate_pct",                "%",   False),
    ("Unsupported Op Count",           "unsupported_op_count",                  "",    True),
    ("CUDA Kernel Count",              "cuda_kernel_count",                     "",    None),
    ("Test Accuracy (%)",              "test_accuracy_pct",                     "%",   False),
    ("Val Accuracy (%)",               "val_accuracy_pct",                      "%",   False),
    ("Test Metric (Hits@50/MRR,%)",   "test_link_metric_pct",                 "%",   False),
    ("Val Metric (Hits@50/MRR,%)",    "val_link_metric_pct",                  "%",   False),
    ("Final Train Loss",               "final_train_loss",                      "",    True),
    ("Required Code Changes",          "required_code_changes",                 "",    True),
    ("Compilation Success",            "compilation_success",                   "",    None),
    ("CPU Util Infer Avg (%)",         "cpu_utilization_inference_pct_avg",     "%",   None),
    ("CPU Util Train Avg (%)",         "cpu_utilization_train_pct_avg",         "%",   None),
    ("Break Even Runs",                "break_even_runs",                       "",    True),
]

# Subset used in summary / median tables
SUMMARY_METRICS = [
    ("Infer Latency Median (ms)",      "inference_latency_median_ms",          "ms",  True),
    ("Infer Latency IQR (ms)",         "inference_latency_iqr_ms",             "ms",  True),
    ("Infer Throughput (nodes/s)",     "throughput_inference_nodes_per_s",     "n/s", False),
    ("Train Throughput (nodes/s)",     "throughput_train_nodes_per_s",         "n/s", False),
    ("Train Throughput (edges/s)",     "throughput_train_edges_per_s",         "e/s", False),
    ("Peak GPU Infer (MB)",            "peak_gpu_memory_inference_mb",         "MB",  True),
    ("Peak GPU Train (MB)",            "peak_gpu_memory_train_mb",             "MB",  True),
    ("Median Epoch Time (s)",          "median_epoch_time_s",                  "s",   True),
    ("Compile Time (s)",               "compile_time_s",                       "s",   True),
    ("Speedup vs Eager",               "speedup_vs_eager",                     "x",   False),
    ("Train Speedup vs Eager",         "train_speedup_vs_eager",               "x",   False),
    ("Graph Capture Rate (%)",         "graph_capture_rate_pct",               "%",   False),
    ("Test Accuracy (%)",              "test_accuracy_pct",                    "%",   False),
    ("Test Metric (Hits@50/MRR,%)",    "test_link_metric_pct",                "%",   False),
]


# Error classification 
def classify_error(err, train_err):
    combined = (err or "") + (train_err or "")
    if not combined.strip():
        return None, None
    if "out of memory" in combined.lower():
        return "OOM", "oom"
    if "DGL forward requires dgl_graph" in combined:
        return "DGL graph err", "dgl_graph"
    if "curr_block->next" in combined:
        return "CUDA mem err", "cuda_mem"
    if "exit 1" in combined:
        return "Crash (exit 1)", "crash"
    return "Error", "other"


# Data loading 
def load_experiments(root_dir):
    """
    Returns dict: (framework, model, dataset) ->
        { mode -> {metric_key: value, '_error': tag, '_error_cat': cat, '_folder': str} }
    """
    root = Path(root_dir)
    subdirs = [d for d in root.iterdir() if d.is_dir()]
    if len(subdirs) == 1 and not any(p.name == "results.json" for p in subdirs[0].iterdir()):
        root = subdirs[0]

    experiments = {}
    for exp_dir in sorted(root.iterdir()):
        if not exp_dir.is_dir():
            continue
        results_file = exp_dir / "results.json"
        config_file  = exp_dir / "config.json"
        if not results_file.exists():
            continue
        try:
            cfg    = json.load(open(config_file))
            config = cfg.get("config", cfg)
            fw, model, dataset = config["framework"], config["model"], config["dataset"]
        except Exception:
            parts   = exp_dir.name.split("_")
            fw      = parts[0] if parts else "unknown"
            model   = parts[1] if len(parts) > 1 else "unknown"
            dataset = "_".join(parts[2:-2]) if len(parts) > 3 else "unknown"

        key = (fw, model, dataset)
        if key not in experiments:
            experiments[key] = {}

        try:
            data    = json.load(open(results_file))
            results = data.get("results", data)
        except Exception:
            continue

        for mode, mdata in results.items():
            err_tag, err_cat = classify_error(mdata.get("error", ""), mdata.get("train_error", ""))
            entry = {"_error": err_tag, "_error_cat": err_cat, "_folder": exp_dir.name}
            for _, key_j, _, _ in METRICS:
                entry[key_j] = mdata.get(key_j)

            # Always read every possible throughput and link-metric key from raw JSON
            entry["throughput_inference_triplets_per_s"] = mdata.get("throughput_inference_triplets_per_s")
            entry["throughput_train_edges_per_s"]        = mdata.get("throughput_train_edges_per_s")
            entry["throughput_train_triplets_per_s"]     = mdata.get("throughput_train_triplets_per_s")
            entry["test_link_metric_pct"]  = mdata.get("test_link_metric_pct")
            entry["val_link_metric_pct"]   = mdata.get("val_link_metric_pct")
            entry["link_metric_name"]      = mdata.get("link_metric_name")
            entry["train_speedup_vs_eager"] = mdata.get("train_speedup_vs_eager")

            # Dataset-type flags
            is_link    = entry["test_link_metric_pct"] is not None
            is_triplet = entry.get("throughput_train_triplets_per_s") is not None
            is_collab  = is_link and not is_triplet  # ogbl-collab / ogbl-citation2
            is_biokg   = is_link and is_triplet       # ogbl-biokg (triplet-based KG)
            entry["_is_link"]    = is_link
            entry["_is_triplet"] = is_triplet

            # Bridge infer throughput for biokg (triplets â†’ nodes/s column is wrong;
            # keep the dedicated triplet column, leave nodes/s as-is / None)
            if is_biokg:
                # infer throughput: use triplets/s in dedicated column only
                if entry.get("throughput_inference_nodes_per_s") is None:
                    entry["throughput_inference_nodes_per_s"] = mdata.get("throughput_inference_triplets_per_s")

            # Bridge train throughput for link datasets into the dedicated columns
            # so they never show â€” :
            #   ogbl-collab  â†’ throughput_train_edges_per_s   (already stored above)
            #   ogbl-biokg   â†’ throughput_train_triplets_per_s (already stored above)
            # The plain nodes/s column stays None for link datasets (correct).

            experiments[key][mode] = entry

    return experiments


def get_val(exps, key, mode, metric_key):
    """Return numeric value or None if missing/errored."""
    entry = exps.get(key, {}).get(mode, {})
    if not entry or entry.get("_error"):
        return None
    v = entry.get(metric_key)
    return v if isinstance(v, (int, float)) else None


def safe_median(vals):
    vals = [v for v in vals if v is not None]
    return statistics.median(vals) if vals else None


def n_valid(vals):
    return sum(1 for v in vals if v is not None)

def safe_merge_exps(exp_dicts):
    """Safely merge experiment dicts ensuring modes are updated, not fully overwritten."""
    merged = {}
    for exps in exp_dicts:
        for k, v in exps.items():
            if k not in merged:
                merged[k] = {}
            merged[k].update(v)
    return merged


# Cell helpers 
def sc(cell, value="", font=None, fill=None, alignment=None):
    cell.value     = value
    cell.font      = font or BODY_FONT
    cell.fill      = fill or NO_FILL
    cell.alignment = alignment or LEFT
    cell.border    = BORDER
    return cell


def hdr(ws, r, c, text, fill=HEADER_FILL, font=HEADER_FONT, merge_to=None):
    cell = ws.cell(r, c, text)
    cell.font = font; cell.fill = fill; cell.alignment = CENTER; cell.border = BORDER
    if merge_to:
        ws.merge_cells(f"{get_column_letter(c)}{r}:{get_column_letter(merge_to)}{r}")
    return cell


def subhdr(ws, r, c, text, fill=SUBHDR_FILL):
    return hdr(ws, r, c, text, fill=fill, font=SUBHDR_FONT)


def data_cell(ws, r, c, val, fill=None, err=False, neutral=False, is_link=False):
    cell = ws.cell(r, c)
    # italic font used to mark link-prediction metrics (Hits@50) and edges/s throughput
    link_font = Font(name="Arial", size=9, italic=True)
    if err:
        cell.value = f"[{val}]" if val else "ERR"
        cell.font = ERROR_FONT; cell.fill = ORANGE_FILL; cell.alignment = CENTER
    elif val is None:
        cell.value = "â€”"; cell.font = BODY_FONT
        cell.fill  = GREY_FILL if neutral else (fill or NO_FILL)
        cell.alignment = CENTER
    elif isinstance(val, bool):
        cell.value = "âœ“" if val else "âœ—"
        cell.font = link_font if is_link else BODY_FONT
        cell.fill = fill or NO_FILL; cell.alignment = CENTER
    elif isinstance(val, float):
        cell.value = round(val, 4)
        cell.font = link_font if is_link else BODY_FONT
        cell.fill = fill or NO_FILL; cell.alignment = RIGHT
    else:
        cell.value = val
        cell.font = link_font if is_link else BODY_FONT
        cell.fill = fill or NO_FILL; cell.alignment = RIGHT
    cell.border = BORDER
    return cell


LINK_ITALIC_KEYS = {
    # These keys carry link-prediction or KG semantics â€” italicise to distinguish
    # from node-classification accuracy and nodes/s throughput.
    # Italic in column headers signals the same thing.
    "test_link_metric_pct",
    "val_link_metric_pct",
    "throughput_train_edges_per_s",
    "throughput_train_triplets_per_s",
    "throughput_inference_triplets_per_s",
    # nodes/s columns bridged from edges/s for biokg infer
    "throughput_inference_nodes_per_s",
    "throughput_train_nodes_per_s",
}

def is_link_metric(entry, key_j):
    """Return True when this entry is a link-prediction result AND the column
    is one that carries a different meaning for link vs node datasets."""
    return bool(entry and entry.get("_is_link") and key_j in LINK_ITALIC_KEYS)


def delta_cell(ws, r, c, base_val, new_val, lower_is_better, row_fill=None):
    cell = ws.cell(r, c)
    if base_val is None or new_val is None:
        cell.value = "â€”"; cell.fill = GREY_FILL
    else:
        try:
            d   = float(new_val) - float(base_val)
            pct = (d / float(base_val) * 100) if float(base_val) != 0 else 0.0
            cell.value = f"{'+'if d>=0 else''}{pct:.1f}%"
            if lower_is_better is None or d == 0:
                cell.fill = row_fill or NO_FILL
            elif (d < 0) == lower_is_better:
                cell.fill = GREEN_FILL
            else:
                cell.fill = RED_FILL
        except Exception:
            cell.value = "â€”"; cell.fill = GREY_FILL
    cell.font = BOLD_FONT; cell.alignment = CENTER; cell.border = BORDER


def cw(ws, mapping):
    """Set column widths. mapping: {col_int_or_letter: width}"""
    for col, w in mapping.items():
        k = get_column_letter(col) if isinstance(col, int) else col
        ws.column_dimensions[k].width = w


def freeze_rows(ws, cell, h1=40, h2=28):
    ws.freeze_panes = cell
    ws.row_dimensions[1].height = h1
    if h2:
        ws.row_dimensions[2].height = h2



# SHEET 0: Research Questions overview

def build_rq_overview_sheet(wb, false_exps, true_exps, auto_exps, nofixes_exps=None):
    ws = wb.create_sheet("Research Questions", 0)

    def title(r, text):
        cell = ws.cell(r, 1, text)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
        cell.fill = HEADER_FILL; cell.alignment = CENTER; cell.border = BORDER
        ws.merge_cells(f"A{r}:F{r}")
        ws.row_dimensions[r].height = 36

    def rq_row(r, text):
        cell = ws.cell(r, 1, text)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = SUBHDR_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.merge_cells(f"A{r}:F{r}")
        ws.row_dimensions[r].height = 30

    def stat_row(r, label, val, good=None):
        lc = ws.cell(r, 1, label)
        lc.font = BOLD_FONT; lc.fill = ALT_ROW_FILL; lc.alignment = LEFT; lc.border = BORDER
        ws.merge_cells(f"A{r}:C{r}")
        vc = ws.cell(r, 4, val)
        vc.font = BODY_FONT
        vc.fill = GREEN_FILL if good is True else (RED_FILL if good is False else NO_FILL)
        vc.alignment = LEFT; vc.border = BORDER
        ws.merge_cells(f"D{r}:F{r}")

    # Aggregate stats across all 4 folders
    _all_exp_sets = [false_exps, true_exps, auto_exps] + ([nofixes_exps] if nofixes_exps else [])
    all_entries = [e for exps in _all_exp_sets for modes in exps.values() for e in modes.values()]
    total = len(all_entries)
    n_oom = sum(1 for e in all_entries if e.get("_error_cat") == "oom")
    n_err = sum(1 for e in all_entries if e.get("_error_cat") and e["_error_cat"] != "oom")
    n_ok  = total - n_oom - n_err

    true_better = true_worse = false_better = false_worse = 0
    for key in set(auto_exps) & set(true_exps):
        for mode in COMPILE_MODES:
            av = get_val(auto_exps,  key, mode, "inference_latency_median_ms")
            tv = get_val(true_exps,  key, mode, "inference_latency_median_ms")
            fv = get_val(false_exps, key, mode, "inference_latency_median_ms")
            if av and tv:
                if tv < av:   true_better += 1
                elif tv > av: true_worse  += 1
            if av and fv:
                if fv < av:   false_better += 1
                elif fv > av: false_worse  += 1

    pair_data = {}
    for exps in [false_exps, true_exps, auto_exps]:
        for (fw, model, ds), modes in exps.items():
            pair_data.setdefault((model, ds), {})[fw] = modes
            
    dgl_faster = pyg_faster = 0
    for fwd in pair_data.values():
        dv = (fwd.get("dgl",{}).get("eager",{}) or {}).get("inference_latency_median_ms")
        pv = (fwd.get("pyg",{}).get("eager",{}) or {}).get("inference_latency_median_ms")
        if dv and pv and not fwd.get("dgl",{}).get("eager",{}).get("_error") \
                     and not fwd.get("pyg",{}).get("eager",{}).get("_error"):
            if dv < pv: dgl_faster += 1
            elif pv < dv: pyg_faster += 1

    r = 1
    title(r, "GNN torch.compile Benchmark â€” Research Questions & Results Summary"); r += 2

    RQs = [
        ("RQ1: Does torch.compile improve inference latency for GNNs (PyG & DGL)?",
         [("Key metrics", "Infer Latency Median, Speedup vs Eager, Train Throughput"),
          ("See sheets",  "Per-mode sheets (Eager, Default, â€¦) â†’ Speedup vs Eager column"),
          ("Summary",     "Mode Summary sheets")]),
        ("RQ2: How does compile overhead (compile time, break-even) affect usability?",
         [("Key metrics", "Compile Time (s), Total Compile Tax, Compile OH (equiv calls), Break Even Runs"),
          ("See sheets",  "Per-mode comparison sheets â€” compile columns")]),
        ("RQ3: Does dynamic=True/False/Auto affect graph breaks and compilability?",
         [("True improves latency vs Auto",   f"{true_better} experimentÃ—mode pairs",  true_better > 0),
          ("True regressions vs Auto",        f"{true_worse} experimentÃ—mode pairs",   true_worse == 0),
          ("False improves latency vs Auto",  f"{false_better} experimentÃ—mode pairs", false_better > 0),
          ("False regressions vs Auto",       f"{false_worse} experimentÃ—mode pairs",  false_worse == 0),
          ("Key metrics", "Graph Capture Rate %, Required Code Changes, Unsupported Op Count"),
          ("See sheets",  "Per-mode sheets â†’ Auto (baseline) vs True / False Î” columns")]),
        ("RQ4: Is DGL eager faster / more memory-efficient than PyG eager?",
         [("DGL eager faster (infer latency)", f"{dgl_faster} / {dgl_faster+pyg_faster} pairs", dgl_faster > pyg_faster),
          ("PyG eager faster",                 f"{pyg_faster} / {dgl_faster+pyg_faster} pairs"),
          ("See sheets", "DGL vs PyG (Eager), DGL vs PyG (All Modes)")]),
        ("RQ5: Can PyG compiled modes match or beat DGL eager performance?",
         [("Key insight", "DGL compiled modes are consistently slower than PyG compiled modes"),
          ("Approach",    "DGL eager used as DGL's practical best; compared to all PyG compile modes"),
          ("See sheets",  "DGL Eager vs PyG Compiled")]),
        ("RQ6: Which compile mode performs best per framework/dataset/model?",
         [("See sheets",      "Mode Summary sheets â€” median per mode per framework"),
          ("Per-framework",   "DGL Summary, PyG Summary â€” full breakdown per experiment")]),
    ]
    for rq_text, stats in RQs:
        rq_row(r, rq_text); r += 1
        for s in stats:
            stat_row(r, s[0], s[1], s[2] if len(s) > 2 else None); r += 1
        r += 1

    rq_row(r, "Overall Run Statistics"); r += 1
    for label, val, good in [
        ("Total experiment Ã— mode entries", total, None),
        ("Successful runs",                 n_ok,  True),
        ("OOM failures",                    n_oom, n_oom == 0),
        ("Other errors",                    n_err, n_err == 0),
    ]:
        stat_row(r, label, val, good); r += 1
    r += 1

    rq_row(r, "Color Legend"); r += 1
    for fill, text in [
        (GREEN_FILL,  "ðŸŸ¢  Better than baseline / Improved"),
        (RED_FILL,    "ðŸ”´  Worse than baseline / Regression"),
        (ORANGE_FILL, "ðŸŸ   OOM or Error â€” excluded from numeric comparisons"),
        (GREY_FILL,   "â¬œ  N/A â€” experiment not present / value unavailable"),
        (SUMM_FILL,   "ðŸŸ¡  Median summary row (yellow)"),
        (SUMM2_FILL,  "ðŸŸ¢  Median summary row â€“ framework subset (light green)"),
        (NO_FILL,     "Italic values = link-prediction metric (Hits@50 for collab, MRR for biokg) or edges/s â€” NOT node classification accuracy or nodes/s"),
    ]:
        c = ws.cell(r, 1, text)
        c.font = BODY_FONT; c.fill = fill; c.alignment = LEFT; c.border = BORDER
        ws.merge_cells(f"A{r}:F{r}"); r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["D"].width = 50
    for col in "BCEF":
        ws.column_dimensions[col].width = 8



# SHEET: Per-mode comparison (No WA baseline vs WA vs No-Dyn) + median rows

def build_comparison_sheet(wb, sheet_name, false_exps, true_exps, auto_exps, nofixes_exps, mode):
    ws = wb.create_sheet(sheet_name)
    freeze_rows(ws, "D3")

    all_keys = sorted(set(false_exps) | set(true_exps) | set(auto_exps) | set(nofixes_exps), key=exp_sort_key)

    for ci, label in enumerate(["Framework", "Model", "Dataset"], 1):
        ws.merge_cells(f"{get_column_letter(ci)}1:{get_column_letter(ci)}2")
        hdr(ws, 1, ci, label)

    col = 4
    for disp, _, unit, _ in METRICS:
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+6)}1")
        hdr(ws, 1, col, f"{disp} ({unit})" if unit else disp)
        subhdr(ws, 2, col,   "Auto")
        subhdr(ws, 2, col+1, "True")
        subhdr(ws, 2, col+2, "False")
        subhdr(ws, 2, col+3, "No-Fixes")
        subhdr(ws, 2, col+4, "Î” (True vs Auto)")
        subhdr(ws, 2, col+5, "Î” (False vs Auto)")
        subhdr(ws, 2, col+6, "Î” (No-Fixes vs Auto)")
        col += 7

    row = 3
    for fw, model, dataset in all_keys:
        key  = (fw, model, dataset)
        base = auto_exps.get(key,    {}).get(mode, {})
        work = true_exps.get(key,    {}).get(mode, {})
        wa_e = false_exps.get(key,   {}).get(mode, {})
        nofx = nofixes_exps.get(key, {}).get(mode, {})
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL

        sc(ws.cell(row, 1), fw.upper(),    fill=fill)
        sc(ws.cell(row, 2), model.upper(), fill=fill)
        sc(ws.cell(row, 3), dataset,       fill=fill)

        col = 4
        for _, key_j, _, lib in METRICS:
            b_err  = base.get("_error") if base else None
            w_err  = work.get("_error") if work else None
            wa_err = wa_e.get("_error") if wa_e else None
            nf_err = nofx.get("_error") if nofx else None

            bv  = base.get(key_j) if base and not b_err else None
            wv  = work.get(key_j) if work and not w_err else None
            wav = wa_e.get(key_j) if wa_e and not wa_err else None
            nfv = nofx.get(key_j) if nofx and not nf_err else None

            data_cell(ws, row, col,   b_err  or bv,  fill=fill, err=bool(b_err),  neutral=not base,  is_link=is_link_metric(base, key_j))
            data_cell(ws, row, col+1, w_err  or wv,  fill=fill, err=bool(w_err),  neutral=not work,  is_link=is_link_metric(work, key_j))
            data_cell(ws, row, col+2, wa_err or wav, fill=fill, err=bool(wa_err), neutral=not wa_e,  is_link=is_link_metric(wa_e,  key_j))
            data_cell(ws, row, col+3, nf_err or nfv, fill=fill, err=bool(nf_err), neutral=not nofx,  is_link=is_link_metric(nofx,  key_j))
            delta_cell(ws, row, col+4, bv,  wv,  lib, row_fill=fill)   # Î” True vs Auto
            delta_cell(ws, row, col+5, bv,  wav, lib, row_fill=fill)   # Î” False vs Auto
            delta_cell(ws, row, col+6, bv,  nfv, lib, row_fill=fill)   # Î” No-Fixes vs Auto
            col += 7
        row += 1

    # Median rows: ALL / DGL / PyG
    for group_label, g_keys in [
        ("MEDIAN â€“ ALL", all_keys),
        ("MEDIAN â€“ DGL", [k for k in all_keys if k[0] == "dgl"]),
        ("MEDIAN â€“ PyG", [k for k in all_keys if k[0] == "pyg"]),
    ]:
        fill = SUMM_FILL if "ALL" in group_label else SUMM2_FILL
        sc(ws.cell(row, 1), group_label, font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "",          font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 3), "",          font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in METRICS:
            am  = safe_median([get_val(auto_exps,    k, mode, key_j) for k in g_keys])
            wm  = safe_median([get_val(true_exps,    k, mode, key_j) for k in g_keys])
            fm  = safe_median([get_val(false_exps,   k, mode, key_j) for k in g_keys])
            nfm = safe_median([get_val(nofixes_exps, k, mode, key_j) for k in g_keys])

            data_cell(ws, row, col,   am,  fill=fill)
            data_cell(ws, row, col+1, wm,  fill=fill)
            data_cell(ws, row, col+2, fm,  fill=fill)
            data_cell(ws, row, col+3, nfm, fill=fill)
            delta_cell(ws, row, col+4, am, wm,  lib, row_fill=fill)   # Î” True vs Auto
            delta_cell(ws, row, col+5, am, fm,  lib, row_fill=fill)   # Î” False vs Auto
            delta_cell(ws, row, col+6, am, nfm, lib, row_fill=fill)   # Î” No-Fixes vs Auto
            col += 7
        row += 1

    cw(ws, {1: 10, 2: 12, 3: 18})
    col = 4
    for _ in METRICS:
        cw(ws, {col: 13, col+1: 13, col+2: 13, col+3: 13, col+4: 12, col+5: 12, col+6: 12}); col += 7



# SHEET: Mode Summary (Median pivot)

def build_mode_summary_sheet(wb, exps, sheet_name):
    ws = wb.create_sheet(sheet_name)
    freeze_rows(ws, "C3")

    all_keys = sorted(exps.keys(), key=exp_sort_key)
    groups = [("ALL", all_keys),
              ("DGL", [k for k in all_keys if k[0] == "dgl"]),
              ("PyG", [k for k in all_keys if k[0] == "pyg"])]

    ws.merge_cells("A1:A2"); hdr(ws, 1, 1, "Compile Mode")
    ws.merge_cells("B1:B2"); hdr(ws, 1, 2, "N valid (ALL)")

    col = 3
    for g_label, _ in groups:
        end = col + len(SUMMARY_METRICS) - 1
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(end)}1")
        fill = HEADER_FILL if g_label == "ALL" else SUBHDR_FILL
        hdr(ws, 1, col, g_label, fill=fill)
        for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
            subhdr(ws, 2, col + i, f"{disp} ({unit})" if unit else disp)
        col += len(SUMMARY_METRICS)

    row = 3
    for mode in COMPILE_MODES:
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        sc(ws.cell(row, 1), MODE_SHORT[mode], font=BOLD_FONT, fill=fill)
        n = n_valid([get_val(exps, k, mode, "inference_latency_median_ms") for k in all_keys])
        sc(ws.cell(row, 2), n, font=BODY_FONT, fill=fill, alignment=CENTER)
        col = 3
        for _, g_keys in groups:
            for _, key_j, _, _ in SUMMARY_METRICS:
                med = safe_median([get_val(exps, k, mode, key_j) for k in g_keys])
                data_cell(ws, row, col, med, fill=fill)
                col += 1
        row += 1

    row += 1
    ws.merge_cells(f"A{row}:B{row}")
    hdr(ws, row, 1, "Î” vs Eager (median)", fill=SUBHDR_FILL)
    row += 1
    for mode in COMPILE_MODES:
        if mode == "eager":
            continue
        fill = SUMM2_FILL
        sc(ws.cell(row, 1), f"Î” {MODE_SHORT[mode]} vs eager", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "", fill=fill)
        col = 3
        for _, g_keys in groups:
            for _, key_j, _, lib in SUMMARY_METRICS:
                eager_med = safe_median([get_val(exps, k, "eager", key_j) for k in g_keys])
                mode_med  = safe_median([get_val(exps, k, mode,   key_j) for k in g_keys])
                delta_cell(ws, row, col, eager_med, mode_med, lib, row_fill=fill)
                col += 1
        row += 1

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    col = 3
    for _ in groups:
        for _ in SUMMARY_METRICS:
            ws.column_dimensions[get_column_letter(col)].width = 14
            col += 1



# SHEET: Per-framework deep dive 

def build_per_framework_summary(wb, false_exps, true_exps, auto_exps, framework, sheet_name, nofixes_exps=None):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "C4"
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 20

    fw_true   = sorted((k for k in true_exps    if k[0] == framework), key=exp_sort_key)
    fw_false  = sorted((k for k in false_exps   if k[0] == framework), key=exp_sort_key)
    fw_auto   = sorted((k for k in auto_exps    if k[0] == framework), key=exp_sort_key)
    fw_nofx   = sorted((k for k in (nofixes_exps or {}) if k[0] == framework), key=exp_sort_key)
    fw_all    = sorted(set(fw_true) | set(fw_false) | set(fw_auto) | set(fw_nofx), key=exp_sort_key)

    ws.merge_cells("A1:A3"); hdr(ws, 1, 1, "Model")
    ws.merge_cells("B1:B3"); hdr(ws, 1, 2, "Dataset")

    col = 3
    _N_PER_METRIC = 7  # Auto | True | False | No-Fixes | Î”(T-A) | Î”(F-A) | Î”(NF-A)
    for mode in COMPILE_MODES:
        n_metrics = len(SUMMARY_METRICS)
        span = n_metrics * _N_PER_METRIC
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+span-1)}1")
        hdr(ws, 1, col, MODE_SHORT[mode])
        for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
            bc = col + i * _N_PER_METRIC
            ws.merge_cells(f"{get_column_letter(bc)}2:{get_column_letter(bc+_N_PER_METRIC-1)}2")
            subhdr(ws, 2, bc, f"{disp} ({unit})" if unit else disp, fill=SUBHDR2_FILL)
            subhdr(ws, 3, bc,   "Auto",              fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+1, "True",              fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+2, "False",             fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+3, "No-Fixes",          fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+4, "Î” (True-Auto)",     fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+5, "Î” (False-Auto)",    fill=SUBHDR_FILL)
            subhdr(ws, 3, bc+6, "Î” (No-Fixes-Auto)", fill=SUBHDR_FILL)
        col += span

    row = 4
    for fw, model, dataset in fw_all:
        key  = (fw, model, dataset)
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        sc(ws.cell(row, 1), model.upper(), fill=fill)
        sc(ws.cell(row, 2), dataset,       fill=fill)
        col = 3
        for mode in COMPILE_MODES:
            for _, key_j, _, lib in SUMMARY_METRICS:
                b_entry  = auto_exps.get(key,          {}).get(mode, {})
                w_entry  = true_exps.get(key,           {}).get(mode, {})
                wa_entry = false_exps.get(key,          {}).get(mode, {})
                nf_entry = (nofixes_exps or {}).get(key,{}).get(mode, {})

                b_err  = b_entry.get("_error")  if b_entry  else None
                w_err  = w_entry.get("_error")  if w_entry  else None
                wa_err = wa_entry.get("_error") if wa_entry else None
                nf_err = nf_entry.get("_error") if nf_entry else None

                bv  = b_entry.get(key_j)  if b_entry  and not b_err  else None
                wv  = w_entry.get(key_j)  if w_entry  and not w_err  else None
                wav = wa_entry.get(key_j) if wa_entry and not wa_err else None
                nfv = nf_entry.get(key_j) if nf_entry and not nf_err else None

                data_cell(ws, row, col,   b_err  or bv,  fill=fill, err=bool(b_err),  neutral=not b_entry,  is_link=is_link_metric(b_entry,  key_j))
                data_cell(ws, row, col+1, w_err  or wv,  fill=fill, err=bool(w_err),  neutral=not w_entry,  is_link=is_link_metric(w_entry,  key_j))
                data_cell(ws, row, col+2, wa_err or wav, fill=fill, err=bool(wa_err), neutral=not wa_entry, is_link=is_link_metric(wa_entry, key_j))
                data_cell(ws, row, col+3, nf_err or nfv, fill=fill, err=bool(nf_err), neutral=not nf_entry, is_link=is_link_metric(nf_entry, key_j))
                delta_cell(ws, row, col+4, bv,  wv,  lib, row_fill=fill)   # Î” True vs Auto
                delta_cell(ws, row, col+5, bv,  wav, lib, row_fill=fill)   # Î” False vs Auto
                delta_cell(ws, row, col+6, bv,  nfv, lib, row_fill=fill)   # Î” No-Fixes vs Auto
                col += _N_PER_METRIC
        row += 1

    models_present = sorted(set(k[1] for k in fw_all), key=_model_rank)
    for group_label, g_keys in (
        [(f"MEDIAN â€“ {framework.upper()} ALL", fw_all)] +
        [(f"MEDIAN â€“ {m.upper()}", [k for k in fw_all if k[1] == m]) for m in models_present]
    ):
        fill = SUMM_FILL
        sc(ws.cell(row, 1), group_label, font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "",          font=SUMM_FONT, fill=fill)
        col = 3
        for mode in COMPILE_MODES:
            for _, key_j, _, lib in SUMMARY_METRICS:
                am  = safe_median([get_val(auto_exps,          k, mode, key_j) for k in g_keys])
                wm  = safe_median([get_val(true_exps,          k, mode, key_j) for k in g_keys])
                fm  = safe_median([get_val(false_exps,         k, mode, key_j) for k in g_keys])
                nfm = safe_median([get_val(nofixes_exps or {}, k, mode, key_j) for k in g_keys])

                data_cell(ws, row, col,   am,  fill=fill)
                data_cell(ws, row, col+1, wm,  fill=fill)
                data_cell(ws, row, col+2, fm,  fill=fill)
                data_cell(ws, row, col+3, nfm, fill=fill)
                delta_cell(ws, row, col+4, am, wm,  lib, row_fill=fill)   # Î” True vs Auto
                delta_cell(ws, row, col+5, am, fm,  lib, row_fill=fill)   # Î” False vs Auto
                delta_cell(ws, row, col+6, am, nfm, lib, row_fill=fill)   # Î” No-Fixes vs Auto
                col += _N_PER_METRIC
        row += 1

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 18
    col = 3
    for _ in COMPILE_MODES:
        for _ in SUMMARY_METRICS:
            cw(ws, {col: 11, col+1: 11, col+2: 11, col+3: 11, col+4: 9, col+5: 9, col+6: 9})
            col += _N_PER_METRIC



# SHEET: DGL vs PyG (Eager) 

def build_fw_eager_sheet(wb, merged_exps):
    ws = wb.create_sheet("DGL vs PyG (Eager)")
    freeze_rows(ws, "C3")

    pairs  = {}
    for (fw, model, ds), modes in merged_exps.items():
        pairs.setdefault((model, ds), {})[fw] = modes
    sorted_pairs = sorted(pairs, key=pair_sort_key)

    ws.merge_cells("A1:A2"); hdr(ws, 1, 1, "Model")
    ws.merge_cells("B1:B2"); hdr(ws, 1, 2, "Dataset")
    col = 3
    for disp, _, unit, _ in METRICS:
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+2)}1")
        hdr(ws, 1, col, f"{disp} ({unit})" if unit else disp)
        subhdr(ws, 2, col,   "DGL eager")
        subhdr(ws, 2, col+1, "PyG eager")
        subhdr(ws, 2, col+2, "Î” (DGLâ†’PyG)")
        col += 3

    row = 3
    for model, ds in sorted_pairs:
        fwd  = pairs[(model, ds)]
        de   = fwd.get("dgl", {}).get("eager", {})
        pe   = fwd.get("pyg", {}).get("eager", {})
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        sc(ws.cell(row, 1), model.upper(), fill=fill)
        sc(ws.cell(row, 2), ds,            fill=fill)
        col = 3
        for _, key_j, _, lib in METRICS:
            de_err = de.get("_error") if de else None
            pe_err = pe.get("_error") if pe else None
            dv = de.get(key_j) if de and not de_err else None
            pv = pe.get(key_j) if pe and not pe_err else None
            data_cell(ws, row, col,   de_err or dv, fill=fill, err=bool(de_err), neutral=not de)
            data_cell(ws, row, col+1, pe_err or pv, fill=fill, err=bool(pe_err), neutral=not pe)
            delta_cell(ws, row, col+2, dv, pv, lib, row_fill=fill)
            col += 3
        row += 1

    for label in ["MEDIAN â€“ ALL"]:
        fill = SUMM_FILL
        sc(ws.cell(row, 1), label, font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "",    font=SUMM_FONT, fill=fill)
        col = 3
        for _, key_j, _, lib in METRICS:
            def _collect(fw_key):
                return [e.get(key_j) for k in sorted_pairs
                        for e in [pairs[k].get(fw_key, {}).get("eager", {})]
                        if e and not e.get("_error") and isinstance(e.get(key_j), (int, float))]
            dm = safe_median(_collect("dgl"))
            pm = safe_median(_collect("pyg"))
            data_cell(ws, row, col,   dm, fill=fill)
            data_cell(ws, row, col+1, pm, fill=fill)
            delta_cell(ws, row, col+2, dm, pm, lib, row_fill=fill)
            col += 3
        row += 1

    cw(ws, {1: 13, 2: 18})
    col = 3
    for _ in METRICS:
        cw(ws, {col: 13, col+1: 13, col+2: 9}); col += 3



# SHEET: DGL vs PyG (All Modes)

def build_fw_all_modes_sheet(wb, merged_exps):
    ws = wb.create_sheet("DGL vs PyG (All Modes)")
    freeze_rows(ws, "D3")

    pairs  = {}
    for (fw, model, ds), modes in merged_exps.items():
        pairs.setdefault((model, ds), {})[fw] = modes
    sorted_pairs = sorted(pairs, key=pair_sort_key)

    ws.merge_cells("A1:A2"); hdr(ws, 1, 1, "Model")
    ws.merge_cells("B1:B2"); hdr(ws, 1, 2, "Dataset")
    ws.merge_cells("C1:C2"); hdr(ws, 1, 3, "Mode")
    col = 4
    for disp, _, unit, _ in METRICS:
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+2)}1")
        hdr(ws, 1, col, f"{disp} ({unit})" if unit else disp)
        subhdr(ws, 2, col,   "DGL")
        subhdr(ws, 2, col+1, "PyG")
        subhdr(ws, 2, col+2, "Î”")
        col += 3

    row = 3
    for mode in COMPILE_MODES:
        for model, ds in sorted_pairs:
            fwd  = pairs[(model, ds)]
            de   = fwd.get("dgl", {}).get(mode, {})
            pe   = fwd.get("pyg", {}).get(mode, {})
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
            sc(ws.cell(row, 1), model.upper(),    fill=fill)
            sc(ws.cell(row, 2), ds,               fill=fill)
            sc(ws.cell(row, 3), MODE_SHORT[mode], fill=fill)
            col = 4
            for _, key_j, _, lib in METRICS:
                de_err = de.get("_error") if de else None
                pe_err = pe.get("_error") if pe else None
                dv = de.get(key_j) if de and not de_err else None
                pv = pe.get(key_j) if pe and not pe_err else None
                data_cell(ws, row, col,   de_err or dv, fill=fill, err=bool(de_err), neutral=not de)
                data_cell(ws, row, col+1, pe_err or pv, fill=fill, err=bool(pe_err), neutral=not pe)
                delta_cell(ws, row, col+2, dv, pv, lib, row_fill=fill)
                col += 3
            row += 1

        fill = SUMM_FILL
        sc(ws.cell(row, 1), f"MEDIAN â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 3), "", font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in METRICS:
            def _col(fw_key, m=mode):
                return [e.get(key_j) for k in sorted_pairs
                        for e in [pairs[k].get(fw_key, {}).get(m, {})]
                        if e and not e.get("_error") and isinstance(e.get(key_j), (int, float))]
            dm = safe_median(_col("dgl"))
            pm = safe_median(_col("pyg"))
            data_cell(ws, row, col,   dm, fill=fill)
            data_cell(ws, row, col+1, pm, fill=fill)
            delta_cell(ws, row, col+2, dm, pm, lib, row_fill=fill)
            col += 3
        row += 2

    cw(ws, {1: 13, 2: 18, 3: 14})
    col = 4
    for _ in METRICS:
        cw(ws, {col: 13, col+1: 13, col+2: 9}); col += 3



# SHEET: DGL Eager vs PyG (Each Mode)

def build_dgl_eager_vs_pyg_compiled_sheet(wb, merged_exps):
    ws = wb.create_sheet("DGL Eager vs PyG Compiled")

    pairs  = {}
    for (fw, model, ds), modes in merged_exps.items():
        pairs.setdefault((model, ds), {})[fw] = modes

    sorted_pairs = sorted(
        (k for k in pairs if pairs[k].get("dgl") or pairs[k].get("pyg")),
        key=pair_sort_key
    )

    n_dgl_cols  = len(SUMMARY_METRICS)          
    n_pyg_cols  = len(SUMMARY_METRICS) * 2      
    n_fixed     = 2                              

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A1:A3"); hdr(ws, 1, 1, "Model")
    ws.merge_cells("B1:B3"); hdr(ws, 1, 2, "Dataset")

    dgl_start = n_fixed + 1
    dgl_end   = dgl_start + n_dgl_cols - 1
    ws.merge_cells(f"{get_column_letter(dgl_start)}1:{get_column_letter(dgl_end)}1")
    hdr(ws, 1, dgl_start, "DGL  eager  (baseline)", fill=PatternFill("solid", start_color="1F4E79"))

    for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
        c = dgl_start + i
        ws.merge_cells(f"{get_column_letter(c)}2:{get_column_letter(c)}3")
        subhdr(ws, 2, c, f"{disp}\n({unit})" if unit else disp, fill=SUBHDR2_FILL)

    col = dgl_end + 1
    MODE_FILLS = {
        "eager":                      PatternFill("solid", start_color="375623"),
        "default":                    PatternFill("solid", start_color="1F4E79"),
        "reduce-overhead":            PatternFill("solid", start_color="833C00"),
        "max-autotune":               PatternFill("solid", start_color="4A235A"),
        "max-autotune-no-cudagraphs": PatternFill("solid", start_color="1A3A4A"),
    }
    for pyg_mode in COMPILE_MODES:
        block_start = col
        block_end   = col + n_pyg_cols - 1
        ws.merge_cells(f"{get_column_letter(block_start)}1:{get_column_letter(block_end)}1")
        hdr(ws, 1, block_start, f"PyG  [{MODE_SHORT[pyg_mode]}]", fill=MODE_FILLS[pyg_mode])

        for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
            val_c   = col + i * 2
            delta_c = val_c + 1
            ws.merge_cells(f"{get_column_letter(val_c)}2:{get_column_letter(delta_c)}2")
            subhdr(ws, 2, val_c, f"{disp}\n({unit})" if unit else disp, fill=SUBHDR2_FILL)
            subhdr(ws, 3, val_c,   "PyG val", fill=SUBHDR_FILL)
            subhdr(ws, 3, delta_c, "Î”",       fill=SUBHDR_FILL)
        col += n_pyg_cols

    ws.freeze_panes = f"{get_column_letter(dgl_start)}4"
    BEST_FONT = Font(name="Arial", bold=True, size=9)

    def get_best_cols_cross(model, ds):
        fwd       = pairs[(model, ds)]
        dgl_entry = fwd.get("dgl", {}).get("eager", {})
        de_err    = dgl_entry.get("_error") if dgl_entry else None
        best = {}
        for i, (_, key_j, _, lower_is_better) in enumerate(SUMMARY_METRICS):
            if lower_is_better is None: continue
            candidates = []
            dv = dgl_entry.get(key_j) if dgl_entry and not de_err else None
            if isinstance(dv, (int, float)):
                candidates.append((dv, dgl_start + i))
            col_scan = dgl_end + 1
            for pyg_mode in COMPILE_MODES:
                pe     = fwd.get("pyg", {}).get(pyg_mode, {})
                pe_err = pe.get("_error") if pe else None
                pv = pe.get(key_j) if pe and not pe_err else None
                if isinstance(pv, (int, float)):
                    candidates.append((pv, col_scan + i * 2))
                col_scan += n_pyg_cols
            if candidates:
                best_val = (min if lower_is_better else max)(c[0] for c in candidates)
                best[i] = {c[1] for c in candidates if c[0] == best_val}
        return best

    row = 4
    for model, ds in sorted_pairs:
        fwd       = pairs[(model, ds)]
        dgl_entry = fwd.get("dgl", {}).get("eager", {})
        de_err    = dgl_entry.get("_error") if dgl_entry else None
        fill      = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        best_cols = get_best_cols_cross(model, ds)

        sc(ws.cell(row, 1), model.upper(), fill=fill)
        sc(ws.cell(row, 2), ds,            fill=fill)

        for i, (_, key_j, _, _) in enumerate(SUMMARY_METRICS):
            dv   = dgl_entry.get(key_j) if dgl_entry and not de_err else None
            c    = dgl_start + i
            cell = data_cell(ws, row, c, de_err or dv, fill=fill,
                             err=bool(de_err), neutral=not dgl_entry)
            if not de_err and dv is not None and i in best_cols and c in best_cols[i]:
                cell.font = BEST_FONT

        col = dgl_end + 1
        for pyg_mode in COMPILE_MODES:
            pe     = fwd.get("pyg", {}).get(pyg_mode, {})
            pe_err = pe.get("_error") if pe else None
            for i, (_, key_j, _, lib) in enumerate(SUMMARY_METRICS):
                val_c   = col + i * 2
                delta_c = val_c + 1
                dv   = dgl_entry.get(key_j) if dgl_entry and not de_err else None
                pv   = pe.get(key_j)        if pe        and not pe_err else None
                cell = data_cell(ws, row, val_c, pe_err or pv, fill=fill,
                                 err=bool(pe_err), neutral=not pe)
                if not pe_err and pv is not None and i in best_cols and val_c in best_cols[i]:
                    cell.font = BEST_FONT
                delta_cell(ws, row, delta_c, dv, pv, lib, row_fill=fill)
            col += n_pyg_cols
        row += 1

    models_present = sorted(set(k[0] for k in sorted_pairs), key=_model_rank)
    summary_groups = [("MEDIAN â€“ ALL", sorted_pairs)] + [
        (f"MEDIAN â€“ {m.upper()}", [k for k in sorted_pairs if k[0] == m])
        for m in models_present
    ]

    for label, g_pairs in summary_groups:
        fill = SUMM_FILL
        sc(ws.cell(row, 1), label, font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "",    font=SUMM_FONT, fill=fill)

        for i, (_, key_j, _, _) in enumerate(SUMMARY_METRICS):
            vals = [pairs[k].get("dgl",{}).get("eager",{}).get(key_j)
                    for k in g_pairs
                    if pairs[k].get("dgl",{}).get("eager",{})
                    and not pairs[k]["dgl"]["eager"].get("_error")
                    and isinstance(pairs[k].get("dgl",{}).get("eager",{}).get(key_j), (int,float))]
            data_cell(ws, row, dgl_start + i, safe_median(vals), fill=fill)

        col = dgl_end + 1
        for pyg_mode in COMPILE_MODES:
            for i, (_, key_j, _, lib) in enumerate(SUMMARY_METRICS):
                val_c   = col + i * 2
                delta_c = val_c + 1
                dgl_vals = [pairs[k].get("dgl",{}).get("eager",{}).get(key_j)
                            for k in g_pairs
                            if pairs[k].get("dgl",{}).get("eager",{})
                            and not pairs[k]["dgl"]["eager"].get("_error")
                            and isinstance(pairs[k].get("dgl",{}).get("eager",{}).get(key_j),(int,float))]
                pyg_vals = [pairs[k].get("pyg",{}).get(pyg_mode,{}).get(key_j)
                            for k in g_pairs
                            if pairs[k].get("pyg",{}).get(pyg_mode,{})
                            and not pairs[k].get("pyg",{}).get(pyg_mode,{}).get("_error")
                            and isinstance(pairs[k].get("pyg",{}).get(pyg_mode,{}).get(key_j),(int,float))]
                dm = safe_median(dgl_vals)
                pm = safe_median(pyg_vals)
                data_cell(ws, row, val_c,   pm, fill=fill)
                delta_cell(ws, row, delta_c, dm, pm, lib, row_fill=fill)
            col += n_pyg_cols
        row += 1

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 18
    for i in range(n_dgl_cols):
        ws.column_dimensions[get_column_letter(dgl_start + i)].width = 13
    col = dgl_end + 1
    for _ in COMPILE_MODES:
        for i in range(len(SUMMARY_METRICS)):
            ws.column_dimensions[get_column_letter(col + i*2)].width   = 13
            ws.column_dimensions[get_column_letter(col + i*2 + 1)].width = 8
        col += n_pyg_cols



# SHEET: Compile Modes vs Eager (within each framework)

def build_compile_vs_eager_sheet(wb, merged_exps):
    ws = wb.create_sheet("Compile Modes vs Eager")
    all_keys = sorted(merged_exps.keys(), key=exp_sort_key)

    COMPILED_MODES = [m for m in COMPILE_MODES if m != "eager"]
    n_eager_cols   = len(SUMMARY_METRICS)          
    n_mode_cols    = len(SUMMARY_METRICS) * 2      
    n_fixed        = 3                              

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A1:A3"); hdr(ws, 1, 1, "Framework")
    ws.merge_cells("B1:B3"); hdr(ws, 1, 2, "Model")
    ws.merge_cells("C1:C3"); hdr(ws, 1, 3, "Dataset")

    eager_start = n_fixed + 1
    eager_end   = eager_start + n_eager_cols - 1
    ws.merge_cells(f"{get_column_letter(eager_start)}1:{get_column_letter(eager_end)}1")
    hdr(ws, 1, eager_start, "eager  (baseline)", fill=PatternFill("solid", start_color="1F4E79"))
    for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
        c = eager_start + i
        ws.merge_cells(f"{get_column_letter(c)}2:{get_column_letter(c)}3")
        subhdr(ws, 2, c, f"{disp}\n({unit})" if unit else disp, fill=SUBHDR2_FILL)

    MODE_FILLS = {
        "default":                    PatternFill("solid", start_color="1F4E79"),
        "reduce-overhead":            PatternFill("solid", start_color="833C00"),
        "max-autotune":               PatternFill("solid", start_color="4A235A"),
        "max-autotune-no-cudagraphs": PatternFill("solid", start_color="1A3A4A"),
    }
    col = eager_end + 1
    for mode in COMPILED_MODES:
        block_end = col + n_mode_cols - 1
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(block_end)}1")
        hdr(ws, 1, col, f"[{MODE_SHORT[mode]}]  vs  eager", fill=MODE_FILLS[mode])
        for i, (disp, _, unit, _) in enumerate(SUMMARY_METRICS):
            val_c   = col + i * 2
            delta_c = val_c + 1
            ws.merge_cells(f"{get_column_letter(val_c)}2:{get_column_letter(delta_c)}2")
            subhdr(ws, 2, val_c, f"{disp}\n({unit})" if unit else disp, fill=SUBHDR2_FILL)
            subhdr(ws, 3, val_c,   "value",      fill=SUBHDR_FILL)
            subhdr(ws, 3, delta_c, "Î” vs eager", fill=SUBHDR_FILL)
        col += n_mode_cols

    ws.freeze_panes = f"{get_column_letter(eager_start)}4"
    BEST_FONT = Font(name="Arial", bold=True, size=9)

    def get_best_cols(key):
        modes_data = merged_exps[key]
        eager_entry = modes_data.get("eager", {})
        e_err = eager_entry.get("_error") if eager_entry else None
        best = {}
        for i, (_, key_j, _, lower_is_better) in enumerate(SUMMARY_METRICS):
            if lower_is_better is None: continue
            candidates = []
            ev = eager_entry.get(key_j) if eager_entry and not e_err else None
            if isinstance(ev, (int, float)):
                candidates.append((ev, eager_start + i))
            col_scan = eager_end + 1
            for mode in COMPILED_MODES:
                m_entry = modes_data.get(mode, {})
                m_err   = m_entry.get("_error") if m_entry else None
                mv = m_entry.get(key_j) if m_entry and not m_err else None
                if isinstance(mv, (int, float)):
                    candidates.append((mv, col_scan + i * 2))
                col_scan += n_mode_cols
            if candidates:
                best_val = (min if lower_is_better else max)(c[0] for c in candidates)
                best[i] = {c[1] for c in candidates if c[0] == best_val}
        return best

    def write_experiment_row(ws, row, key, fill):
        fw, model, ds = key
        modes_data = merged_exps[key]
        eager_entry = modes_data.get("eager", {})
        e_err       = eager_entry.get("_error") if eager_entry else None
        best_cols   = get_best_cols(key)

        sc(ws.cell(row, 1), fw.upper(),    fill=fill)
        sc(ws.cell(row, 2), model.upper(), fill=fill)
        sc(ws.cell(row, 3), ds,            fill=fill)

        for i, (_, key_j, _, _) in enumerate(SUMMARY_METRICS):
            ev  = eager_entry.get(key_j) if eager_entry and not e_err else None
            c   = eager_start + i
            cell = data_cell(ws, row, c, e_err or ev, fill=fill,
                             err=bool(e_err), neutral=not eager_entry)
            if not e_err and ev is not None and i in best_cols and c in best_cols[i]:
                cell.font = BEST_FONT

        col = eager_end + 1
        for mode in COMPILED_MODES:
            m_entry = modes_data.get(mode, {})
            m_err   = m_entry.get("_error") if m_entry else None
            for i, (_, key_j, _, lib) in enumerate(SUMMARY_METRICS):
                val_c   = col + i * 2
                delta_c = val_c + 1
                ev   = eager_entry.get(key_j) if eager_entry and not e_err else None
                mv   = m_entry.get(key_j)     if m_entry    and not m_err  else None
                cell = data_cell(ws, row, val_c, m_err or mv, fill=fill,
                                 err=bool(m_err), neutral=not m_entry)
                if not m_err and mv is not None and i in best_cols and val_c in best_cols[i]:
                    cell.font = BEST_FONT
                delta_cell(ws, row, delta_c, ev, mv, lib, row_fill=fill)
            col += n_mode_cols

    row = 4
    prev_fw = None
    for key in all_keys:
        fw = key[0]
        if prev_fw and fw != prev_fw: row += 1
        prev_fw = fw
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        write_experiment_row(ws, row, key, fill)
        row += 1

    def write_median_row(ws, row, label, g_keys, fill=SUMM_FILL):
        sc(ws.cell(row, 1), label, font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "",    font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 3), "",    font=SUMM_FONT, fill=fill)

        for i, (_, key_j, _, _) in enumerate(SUMMARY_METRICS):
            vals = [merged_exps[k].get("eager",{}).get(key_j) for k in g_keys
                    if merged_exps[k].get("eager",{}) and not merged_exps[k]["eager"].get("_error")
                    and isinstance(merged_exps[k].get("eager",{}).get(key_j), (int, float))]
            data_cell(ws, row, eager_start + i, safe_median(vals), fill=fill)

        col = eager_end + 1
        for mode in COMPILED_MODES:
            for i, (_, key_j, _, lib) in enumerate(SUMMARY_METRICS):
                val_c   = col + i * 2
                delta_c = val_c + 1
                e_vals = [merged_exps[k].get("eager",{}).get(key_j) for k in g_keys
                          if merged_exps[k].get("eager",{}) and not merged_exps[k]["eager"].get("_error")
                          and isinstance(merged_exps[k].get("eager",{}).get(key_j), (int, float))]
                m_vals = [merged_exps[k].get(mode,{}).get(key_j) for k in g_keys
                          if merged_exps[k].get(mode,{}) and not merged_exps[k].get(mode,{}).get("_error")
                          and isinstance(merged_exps[k].get(mode,{}).get(key_j), (int, float))]
                em = safe_median(e_vals)
                mm = safe_median(m_vals)
                data_cell(ws, row, val_c,   mm, fill=fill)
                delta_cell(ws, row, delta_c, em, mm, lib, row_fill=fill)
            col += n_mode_cols

    row += 1
    models_present = sorted(set(k[1] for k in all_keys), key=_model_rank)
    summary_groups = (
        [("MEDIAN â€“ ALL", all_keys, SUMM_FILL),
         ("MEDIAN â€“ DGL", [k for k in all_keys if k[0] == "dgl"], SUMM2_FILL),
         ("MEDIAN â€“ PyG", [k for k in all_keys if k[0] == "pyg"], SUMM2_FILL)] +
        [(f"MEDIAN â€“ {m.upper()}", [k for k in all_keys if k[1] == m], SUMM_FILL)
         for m in models_present]
    )
    for label, g_keys, fill in summary_groups:
        if not g_keys: continue
        write_median_row(ws, row, label, g_keys, fill)
        row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 18
    for i in range(n_eager_cols):
        ws.column_dimensions[get_column_letter(eager_start + i)].width = 13
    col = eager_end + 1
    for _ in COMPILED_MODES:
        for i in range(len(SUMMARY_METRICS)):
            ws.column_dimensions[get_column_letter(col + i*2)].width     = 13
            ws.column_dimensions[get_column_letter(col + i*2 + 1)].width = 9
        col += n_mode_cols



# SHEET: Error Summary  (side-by-side: Auto | True | False per experiment+mode)

def build_error_summary_sheet(wb, false_exps, true_exps, auto_exps, nofixes_exps=None):
    ws = wb.create_sheet("Error Summary")
    ws.freeze_panes = "E3"
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 22

    N_FIXED = 4  # FW | Model | Dataset | Mode
    _cond_list = [("Auto", auto_exps), ("True", true_exps), ("False", false_exps)]
    if nofixes_exps:
        _cond_list.append(("No-Fixes", nofixes_exps))
    CONDITIONS = _cond_list
    N_COND_COLS = 2  # Error Type, Folder

    # header rows
    for ci, label in enumerate(["Framework", "Model", "Dataset", "Mode"], 1):
        ws.merge_cells(f"{get_column_letter(ci)}1:{get_column_letter(ci)}2")
        hdr(ws, 1, ci, label)

    COND_FILLS = {
        "Auto":     PatternFill("solid", start_color="1F4E79"),
        "True":     PatternFill("solid", start_color="375623"),
        "False":    PatternFill("solid", start_color="833C00"),
        "No-Fixes": PatternFill("solid", start_color="4A235A"),
    }
    col = N_FIXED + 1
    for cond, _ in CONDITIONS:
        end = col + N_COND_COLS - 1
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(end)}1")
        hdr(ws, 1, col, f"dynamic={cond}", fill=COND_FILLS[cond])
        subhdr(ws, 2, col,   "Error Type", fill=SUBHDR_FILL)
        subhdr(ws, 2, col+1, "Folder",     fill=SUBHDR_FILL)
        col += N_COND_COLS

    # collect all (fw, model, ds, mode) combos that have ANY error
    all_exp_modes = set()
    for _, exps in CONDITIONS:
        for (fw, model, ds), modes in exps.items():
            for mode in COMPILE_MODES:
                entry = modes.get(mode, {})
                if entry and entry.get("_error"):
                    all_exp_modes.add((fw, model, ds, mode))

    if not all_exp_modes:
        ws.cell(3, 1, "No errors found").font = BODY_FONT
        cw(ws, {1: 14, 2: 13, 3: 20, 4: 26})
        return

    sorted_combos = sorted(all_exp_modes,
        key=lambda x: (*exp_sort_key((x[0],x[1],x[2])), COMPILE_MODES.index(x[3])))

    row = 3
    for fw, model, ds, mode in sorted_combos:
        fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        sc(ws.cell(row, 1), fw.upper(),    fill=fill)
        sc(ws.cell(row, 2), model.upper(), fill=fill)
        sc(ws.cell(row, 3), ds,            fill=fill)
        sc(ws.cell(row, 4), mode,          fill=fill)

        col = N_FIXED + 1
        for cond, exps in CONDITIONS:
            entry  = exps.get((fw, model, ds), {}).get(mode, {})
            err    = entry.get("_error") if entry else None
            folder = entry.get("_folder", "") if entry else ""
            if err:
                cf = ORANGE_FILL; ef = ERROR_FONT
            elif entry:
                cf = fill; ef = BODY_FONT
            else:
                cf = GREY_FILL; ef = BODY_FONT
            for ci2, val in enumerate([err or ("OK" if entry else "â€”"), folder or "â€”"], col):
                c = ws.cell(row, ci2, val)
                c.font = ef; c.fill = cf; c.alignment = LEFT; c.border = BORDER
            col += N_COND_COLS
        row += 1

    # summary counts per condition
    row += 1
    last_col = get_column_letter(N_FIXED + len(CONDITIONS) * N_COND_COLS)
    for cond, exps in CONDITIONS:
        n_err = sum(1 for (fw, model, ds, mode) in sorted_combos
                    if exps.get((fw, model, ds), {}).get(mode, {}).get("_error"))
        n_oom = sum(1 for (fw, model, ds, mode) in sorted_combos
                    if exps.get((fw, model, ds), {}).get(mode, {}).get("_error_cat") == "oom")
        cell = ws.cell(row, 1, f"dynamic={cond}: {n_err} errors  ({n_oom} OOM)")
        cell.font = BOLD_FONT
        cell.fill = ORANGE_FILL if n_err else GREEN_FILL
        cell.alignment = LEFT; cell.border = BORDER
        ws.merge_cells(f"A{row}:{last_col}{row}")
        row += 1

    cw(ws, {1: 11, 2: 13, 3: 20, 4: 26})
    col = N_FIXED + 1
    for _ in CONDITIONS:
        cw(ws, {col: 18, col + 1: 45})
        col += N_COND_COLS



# SHEET: Auto vs No-Fixes  (dynamic=auto as baseline vs no code workarounds)

def build_auto_vs_nofixes_sheet(wb, auto_exps, nofixes_exps):
    """Side-by-side comparison of dynamic=auto vs no-fixes across all modes."""
    ws = wb.create_sheet("Auto vs No-Fixes")
    freeze_rows(ws, "D3")

    all_keys = sorted(set(auto_exps) | set(nofixes_exps), key=exp_sort_key)

    for ci, label in enumerate(["Framework", "Model", "Dataset"], 1):
        ws.merge_cells(f"{get_column_letter(ci)}1:{get_column_letter(ci)}2")
        hdr(ws, 1, ci, label)

    NF_FILL = PatternFill("solid", start_color="4A235A")  # dark purple for No-Fixes header
    col = 4
    for disp, _, unit, _ in METRICS:
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+3)}1")
        hdr(ws, 1, col, f"{disp} ({unit})" if unit else disp)
        subhdr(ws, 2, col,   "Auto",              fill=SUBHDR_FILL)
        subhdr(ws, 2, col+1, "No-Fixes",          fill=NF_FILL)
        subhdr(ws, 2, col+2, "Î” (Auto vs NF)",    fill=SUBHDR_FILL)
        subhdr(ws, 2, col+3, "Speedup Autoâ†’NF",   fill=SUBHDR_FILL)
        col += 4

    COMPILED_MODES = [m for m in COMPILE_MODES if m != "eager"]

    row = 3
    for mode in COMPILE_MODES:
        # section divider row
        divider_col = get_column_letter(4 + len(METRICS) * 4 - 1)
        ws.merge_cells(f"A{row}:{divider_col}{row}")
        cell = ws.cell(row, 1, f"â”€â”€ {MODE_SHORT[mode]} â”€â”€")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = SUBHDR2_FILL; cell.alignment = CENTER; cell.border = BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        for fw, model, dataset in all_keys:
            key  = (fw, model, dataset)
            ae   = auto_exps.get(key,    {}).get(mode, {})
            nfe  = nofixes_exps.get(key, {}).get(mode, {})
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL

            sc(ws.cell(row, 1), fw.upper(),    fill=fill)
            sc(ws.cell(row, 2), model.upper(), fill=fill)
            sc(ws.cell(row, 3), dataset,       fill=fill)

            col = 4
            for _, key_j, _, lib in METRICS:
                a_err  = ae.get("_error")  if ae  else None
                nf_err = nfe.get("_error") if nfe else None
                av  = ae.get(key_j)  if ae  and not a_err  else None
                nfv = nfe.get(key_j) if nfe and not nf_err else None

                data_cell(ws, row, col,   a_err  or av,  fill=fill, err=bool(a_err),  neutral=not ae,  is_link=is_link_metric(ae,  key_j))
                data_cell(ws, row, col+1, nf_err or nfv, fill=fill, err=bool(nf_err), neutral=not nfe, is_link=is_link_metric(nfe, key_j))
                delta_cell(ws, row, col+2, nfv, av, lib, row_fill=fill)   # Î” Auto vs No-Fixes (NF=base)

                # speedup cell: av / nfv  (how much auto beats nofixes)
                speed_cell = ws.cell(row, col+3)
                if isinstance(av,(int,float)) and isinstance(nfv,(int,float)) and av != 0 and nfv != 0 and lib is not None:
                    ratio = (nfv / av) if lib else (av / nfv)  # lower_is_better â†’ NF/auto; higherâ†’ auto/NF
                    speed_cell.value = round(ratio, 3)
                    speed_cell.font  = BODY_FONT
                    speed_cell.fill  = GREEN_FILL if ratio > 1 else (RED_FILL if ratio < 1 else fill)
                    speed_cell.alignment = RIGHT
                else:
                    speed_cell.value = "â€”"; speed_cell.fill = GREY_FILL; speed_cell.alignment = CENTER
                    speed_cell.font = BODY_FONT
                speed_cell.border = BORDER
                col += 4
            row += 1

        # median row for this mode
        fill = SUMM_FILL
        sc(ws.cell(row, 1), f"MEDIAN â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 3), "", font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in METRICS:
            am  = safe_median([get_val(auto_exps,    k, mode, key_j) for k in all_keys])
            nfm = safe_median([get_val(nofixes_exps, k, mode, key_j) for k in all_keys])
            data_cell(ws, row, col,   am,  fill=fill)
            data_cell(ws, row, col+1, nfm, fill=fill)
            delta_cell(ws, row, col+2, nfm, am, lib, row_fill=fill)
            ws.cell(row, col+3).value = ""; ws.cell(row, col+3).fill = fill; ws.cell(row, col+3).border = BORDER
            col += 4
        row += 2

    cw(ws, {1: 10, 2: 12, 3: 18})
    col = 4
    for _ in METRICS:
        cw(ws, {col: 13, col+1: 13, col+2: 11, col+3: 11}); col += 4



# MEMORY SHEETS  (5 separate sheets, one per section)

def _mem_pairs(merged_exps):
    pairs = {}
    for (fw, model, ds), modes in merged_exps.items():
        pairs.setdefault((model, ds), {})[fw] = modes
    return sorted(pairs, key=pair_sort_key), pairs

MEM_INFER  = "peak_gpu_memory_inference_mb"
MEM_TRAIN  = "peak_gpu_memory_train_mb"
MEM_METRICS = [
    ("Peak GPU Infer (MB)", MEM_INFER, "MB", True),
    ("Peak GPU Train (MB)", MEM_TRAIN, "MB", True),
]
PURPLE_FILL_MEM = PatternFill("solid", start_color="E2C4F0")

def _mem_block_hdr(ws, r, c, col_a, col_b, col_d="Î”"):
    for mi, (disp, _, _, _) in enumerate(MEM_METRICS):
        bc = c + mi * 3
        ws.merge_cells(f"{get_column_letter(bc)}{r}:{get_column_letter(bc+2)}{r}")
        subhdr(ws, r,   bc,   disp,  fill=SUBHDR2_FILL)
        subhdr(ws, r+1, bc,   col_a, fill=SUBHDR_FILL)
        subhdr(ws, r+1, bc+1, col_b, fill=SUBHDR_FILL)
        subhdr(ws, r+1, bc+2, col_d, fill=SUBHDR_FILL)

def _set_mem_cw(ws, fixed):
    for col, w in fixed.items():
        ws.column_dimensions[get_column_letter(col) if isinstance(col,int) else col].width = w
    start = max(fixed.keys()) + 1
    for i in range(len(MEM_METRICS) * 3):
        ws.column_dimensions[get_column_letter(start + i)].width = 13 if (i % 3) != 2 else 9


def build_memory_dgl_vs_pyg_sheet(wb, merged_exps):
    sorted_pairs, pairs = _mem_pairs(merged_exps)
    ws = wb.create_sheet("Mem 1 DGL vs PyG")
    freeze_rows(ws, "D3")
    hdr(ws, 1, 1, "Model");   ws.merge_cells("A1:A2")
    hdr(ws, 1, 2, "Dataset"); ws.merge_cells("B1:B2")
    hdr(ws, 1, 3, "Mode");    ws.merge_cells("C1:C2")
    _mem_block_hdr(ws, 1, 4, "DGL", "PyG", "Î” (DGLâ†’PyG)")
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20
    row = 3
    for mode in COMPILE_MODES:
        for model, ds in sorted_pairs:
            fwd  = pairs[(model, ds)]
            de   = fwd.get("dgl", {}).get(mode, {})
            pe   = fwd.get("pyg", {}).get(mode, {})
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
            sc(ws.cell(row, 1), model.upper(),    fill=fill)
            sc(ws.cell(row, 2), ds,               fill=fill)
            sc(ws.cell(row, 3), MODE_SHORT[mode], fill=fill)
            col = 4
            for _, key_j, _, lib in MEM_METRICS:
                de_err = de.get("_error") if de else None
                pe_err = pe.get("_error") if pe else None
                dv = de.get(key_j) if de and not de_err else None
                pv = pe.get(key_j) if pe and not pe_err else None
                data_cell(ws, row, col,   de_err or dv, fill=fill, err=bool(de_err), neutral=not de)
                data_cell(ws, row, col+1, pe_err or pv, fill=fill, err=bool(pe_err), neutral=not pe)
                delta_cell(ws, row, col+2, dv, pv, lib, row_fill=fill)
                col += 3
            row += 1
        fill = SUMM_FILL
        sc(ws.cell(row,1), f"MEDIAN â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row,2), "", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row,3), "", font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in MEM_METRICS:
            dv_a = [pairs[k]["dgl"][mode].get(key_j) for k in sorted_pairs
                    if pairs[k].get("dgl",{}).get(mode) and not pairs[k]["dgl"][mode].get("_error")
                    and isinstance(pairs[k]["dgl"][mode].get(key_j),(int,float))]
            pv_a = [pairs[k]["pyg"][mode].get(key_j) for k in sorted_pairs
                    if pairs[k].get("pyg",{}).get(mode) and not pairs[k]["pyg"][mode].get("_error")
                    and isinstance(pairs[k]["pyg"][mode].get(key_j),(int,float))]
            dm = safe_median(dv_a); pm = safe_median(pv_a)
            data_cell(ws, row, col,   dm, fill=fill)
            data_cell(ws, row, col+1, pm, fill=fill)
            delta_cell(ws, row, col+2, dm, pm, lib, row_fill=fill)
            col += 3
        row += 2
    _set_mem_cw(ws, {1:13, 2:20, 3:14})


def build_memory_per_mode_vs_eager_sheet(wb, merged_exps):
    sorted_pairs, pairs = _mem_pairs(merged_exps)
    ws = wb.create_sheet("Mem 2 Mode vs Eager")
    freeze_rows(ws, "E3")
    hdr(ws, 1, 1, "FW");      ws.merge_cells("A1:A2")
    hdr(ws, 1, 2, "Model");   ws.merge_cells("B1:B2")
    hdr(ws, 1, 3, "Dataset"); ws.merge_cells("C1:C2")
    hdr(ws, 1, 4, "Mode");    ws.merge_cells("D1:D2")
    _mem_block_hdr(ws, 1, 5, "Eager", "Compiled", "Î” (eagerâ†’compiled)")
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20
    row = 3
    COMPILED_MODES = [m for m in COMPILE_MODES if m != "eager"]
    for fw in ("dgl", "pyg"):
        for mode in COMPILED_MODES:
            for model, ds in sorted_pairs:
                fwd     = pairs[(model, ds)]
                eager_e = fwd.get(fw, {}).get("eager", {})
                mode_e  = fwd.get(fw, {}).get(mode, {})
                if not eager_e and not mode_e: continue
                fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
                sc(ws.cell(row,1), fw.upper(),       fill=fill)
                sc(ws.cell(row,2), model.upper(),    fill=fill)
                sc(ws.cell(row,3), ds,               fill=fill)
                sc(ws.cell(row,4), MODE_SHORT[mode], fill=fill)
                col = 5
                for _, key_j, _, lib in MEM_METRICS:
                    ee_err = eager_e.get("_error") if eager_e else None
                    me_err = mode_e.get("_error")  if mode_e  else None
                    ev = eager_e.get(key_j) if eager_e and not ee_err else None
                    mv = mode_e.get(key_j)  if mode_e  and not me_err else None
                    data_cell(ws, row, col,   ee_err or ev, fill=fill, err=bool(ee_err), neutral=not eager_e)
                    data_cell(ws, row, col+1, me_err or mv, fill=fill, err=bool(me_err), neutral=not mode_e)
                    delta_cell(ws, row, col+2, ev, mv, lib, row_fill=fill)
                    col += 3
                row += 1
            fill = SUMM2_FILL
            sc(ws.cell(row,1), fw.upper(),                      font=SUMM_FONT, fill=fill)
            sc(ws.cell(row,2), f"MEDIAN â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
            sc(ws.cell(row,3), "", font=SUMM_FONT, fill=fill)
            sc(ws.cell(row,4), "", font=SUMM_FONT, fill=fill)
            col = 5
            for _, key_j, _, lib in MEM_METRICS:
                def _v(m, f=fw, kj=key_j):
                    return [pairs[k].get(f,{}).get(m,{}).get(kj) for k in sorted_pairs
                            if pairs[k].get(f,{}).get(m,{}) and not pairs[k][f][m].get("_error")
                            and isinstance(pairs[k].get(f,{}).get(m,{}).get(kj),(int,float))]
                em = safe_median(_v("eager")); mm = safe_median(_v(mode))
                data_cell(ws, row, col,   em, fill=fill)
                data_cell(ws, row, col+1, mm, fill=fill)
                delta_cell(ws, row, col+2, em, mm, lib, row_fill=fill)
                col += 3
            row += 2
    _set_mem_cw(ws, {1:10, 2:13, 3:20, 4:14})


def _build_fw_mem_sheet(wb, merged_exps, fw, sheet_name):
    sorted_pairs, pairs = _mem_pairs(merged_exps)
    ws = wb.create_sheet(sheet_name)
    freeze_rows(ws, "D3")
    hdr(ws, 1, 1, "Model");   ws.merge_cells("A1:A2")
    hdr(ws, 1, 2, "Dataset"); ws.merge_cells("B1:B2")
    hdr(ws, 1, 3, "Mode");    ws.merge_cells("C1:C2")
    label = fw.upper()
    _mem_block_hdr(ws, 1, 4, f"{label} Eager", f"{label} Compiled", "Î”")
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20
    row = 3
    COMPILED_MODES = [m for m in COMPILE_MODES if m != "eager"]
    for mode in COMPILED_MODES:
        for model, ds in sorted_pairs:
            fwd     = pairs[(model, ds)]
            eager_e = fwd.get(fw, {}).get("eager", {})
            mode_e  = fwd.get(fw, {}).get(mode, {})
            if not eager_e and not mode_e: continue
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
            sc(ws.cell(row,1), model.upper(),    fill=fill)
            sc(ws.cell(row,2), ds,               fill=fill)
            sc(ws.cell(row,3), MODE_SHORT[mode], fill=fill)
            col = 4
            for _, key_j, _, lib in MEM_METRICS:
                ee_err = eager_e.get("_error") if eager_e else None
                me_err = mode_e.get("_error")  if mode_e  else None
                ev = eager_e.get(key_j) if eager_e and not ee_err else None
                mv = mode_e.get(key_j)  if mode_e  and not me_err else None
                data_cell(ws, row, col,   ee_err or ev, fill=fill, err=bool(ee_err), neutral=not eager_e)
                data_cell(ws, row, col+1, me_err or mv, fill=fill, err=bool(me_err), neutral=not mode_e)
                delta_cell(ws, row, col+2, ev, mv, lib, row_fill=fill)
                col += 3
            row += 1
        fill = SUMM2_FILL
        sc(ws.cell(row,1), f"MEDIAN {label} â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row,2), "", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row,3), "", font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in MEM_METRICS:
            ev_a = [pairs[k].get(fw,{}).get("eager",{}).get(key_j) for k in sorted_pairs
                    if pairs[k].get(fw,{}).get("eager",{}) and not pairs[k][fw]["eager"].get("_error")
                    and isinstance(pairs[k].get(fw,{}).get("eager",{}).get(key_j),(int,float))]
            mv_a = [pairs[k].get(fw,{}).get(mode,{}).get(key_j) for k in sorted_pairs
                    if pairs[k].get(fw,{}).get(mode,{}) and not pairs[k][fw][mode].get("_error")
                    and isinstance(pairs[k].get(fw,{}).get(mode,{}).get(key_j),(int,float))]
            em = safe_median(ev_a); mm = safe_median(mv_a)
            data_cell(ws, row, col,   em, fill=fill)
            data_cell(ws, row, col+1, mm, fill=fill)
            delta_cell(ws, row, col+2, em, mm, lib, row_fill=fill)
            col += 3
        row += 2
    _set_mem_cw(ws, {1:13, 2:20, 3:14})

def build_memory_pyg_sheet(wb, merged_exps):
    _build_fw_mem_sheet(wb, merged_exps, "pyg", "Mem 3 PyG")

def build_memory_dgl_sheet(wb, merged_exps):
    _build_fw_mem_sheet(wb, merged_exps, "dgl", "Mem 4 DGL")


def build_memory_best_sheet(wb, merged_exps):
    sorted_pairs, pairs = _mem_pairs(merged_exps)
    GREEN_FILL2 = PatternFill("solid", start_color="C6EFCE")
    ws = wb.create_sheet("Mem 5 Best Mode")
    freeze_rows(ws, "C3")
    hdr(ws, 1, 1, "Model");   ws.merge_cells("A1:A2")
    hdr(ws, 1, 2, "Dataset"); ws.merge_cells("B1:B2")
    col = 3
    for disp, _, _, _ in MEM_METRICS:
        ws.merge_cells(f"{get_column_letter(col)}1:{get_column_letter(col+4)}1")
        subhdr(ws, 1, col, disp, fill=SUBHDR2_FILL)
        for ci2, lbl in zip(range(col, col+5), ["DGL best (MB)","DGL mode","PyG best (MB)","PyG mode","Î” (DGLâ†’PyG)"]):
            subhdr(ws, 2, ci2, lbl, fill=SUBHDR_FILL)
        col += 5
    ws.row_dimensions[1].height = 22; ws.row_dimensions[2].height = 20

    def best_fw_mem(fwd, fw, key_j):
        bv, bm = None, None
        for m in COMPILE_MODES:
            e = fwd.get(fw, {}).get(m, {})
            if not e or e.get("_error"): continue
            v = e.get(key_j)
            if isinstance(v,(int,float)) and (bv is None or v < bv):
                bv, bm = v, MODE_SHORT[m]
        return bv, bm

    row = 3
    for model, ds in sorted_pairs:
        fwd = pairs[(model, ds)]
        row_data = {kj: (*best_fw_mem(fwd,"dgl",kj), *best_fw_mem(fwd,"pyg",kj))
                    for _, kj, _, _ in MEM_METRICS}
        dv_i, _, pv_i, _ = row_data[MEM_INFER]
        if isinstance(dv_i,(int,float)) and isinstance(pv_i,(int,float)):
            fill = PURPLE_FILL_MEM if dv_i < pv_i else (GREEN_FILL2 if pv_i < dv_i else (ALT_ROW_FILL if row%2==0 else NO_FILL))
        else:
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL
        sc(ws.cell(row,1), model.upper(), fill=fill)
        sc(ws.cell(row,2), ds,            fill=fill)
        col = 3
        for _, kj, _, lib in MEM_METRICS:
            dv, dm, pv, pm = row_data[kj]
            data_cell(ws, row, col,   dv, fill=fill, neutral=(dv is None))
            sc(ws.cell(row, col+1), dm or "â€”", fill=fill, alignment=CENTER)
            data_cell(ws, row, col+2, pv, fill=fill, neutral=(pv is None))
            sc(ws.cell(row, col+3), pm or "â€”", fill=fill, alignment=CENTER)
            delta_cell(ws, row, col+4, dv, pv, lib, row_fill=fill)
            col += 5
        row += 1

    row += 1
    for fill, label in [(PURPLE_FILL_MEM, "DGL best mode < PyG best mode  (purple = DGL wins)"),
                        (GREEN_FILL2,     "PyG best mode < DGL best mode  (green = PyG wins)")]:
        c = ws.cell(row, 1, label)
        c.font = BODY_FONT; c.fill = fill; c.alignment = LEFT; c.border = BORDER
        ws.merge_cells(f"A{row}:F{row}"); row += 1

    cw(ws, {1:13, 2:20})
    col = 3
    for _ in MEM_METRICS:
        cw(ws, {col:13, col+1:12, col+2:13, col+3:12, col+4:9}); col += 5



# SHEET: GIN 300 Epochs vs 20 Epochs
# Compares results_gin_300epochs (accuracy run) against the 20-epoch auto

GIN_ACC_METRICS = [
    # (display, json_key, unit, lower_is_better)
    ("Test Accuracy (%)",          "test_accuracy_pct",                  "%",   False),
    ("Val Accuracy (%)",           "val_accuracy_pct",                   "%",   False),
    ("Test Metric (Hits@50,%)",   "test_link_metric_pct",               "%",   False),
    ("Val Metric (Hits@50,%)",    "val_link_metric_pct",                "%",   False),
    ("Final Train Loss",           "final_train_loss",                   "",    True),
    ("Infer Latency Median (ms)",  "inference_latency_median_ms",        "ms",  True),
    ("Speedup vs Eager",           "speedup_vs_eager",                   "x",   False),
    ("Train Speedup vs Eager",     "train_speedup_vs_eager",             "x",   False),
    ("Median Epoch Time (s)",      "median_epoch_time_s",                "s",   True),
    ("Train Throughput (nodes/s)", "throughput_train_nodes_per_s",       "n/s", False),
    ("Train Throughput (edges/s)","throughput_train_edges_per_s",       "e/s", False),
    ("Peak GPU Infer (MB)",        "peak_gpu_memory_inference_mb",       "MB",  True),
    ("Peak GPU Train (MB)",        "peak_gpu_memory_train_mb",           "MB",  True),
    ("Compile Time (s)",           "compile_time_s",                     "s",   True),
    ("Break Even Runs",            "break_even_runs",                    "",    True),
    ("Graph Capture Rate (%)",     "graph_capture_rate_pct",             "%",   False),
]

EPOCH300_FILL = PatternFill("solid", start_color="833C00")  # dark amber for 300-epoch header


def build_gin_accuracy_sheet(wb, auto_exps, gin300_exps):
    """Compare GIN 20-epoch (auto) vs GIN 300-epoch side-by-side per mode."""
    ws = wb.create_sheet("GIN 300 vs 20 Epochs")
    freeze_rows(ws, "D3")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # Filter to GIN experiments only from both sets
    gin_auto_keys = sorted((k for k in auto_exps  if k[1] == "gin" and k[2] == "ogbn-arxiv"), key=exp_sort_key)
    gin_300_keys  = sorted((k for k in gin300_exps if k[1] == "gin" and k[2] == "ogbn-arxiv"), key=exp_sort_key)
    all_gin_keys  = sorted(set(gin_auto_keys) | set(gin_300_keys), key=exp_sort_key)

    if not all_gin_keys:
        ws.cell(1, 1, "No GIN experiments found in either dataset.").font = BODY_FONT
        return

    # columns: FW | Model | Dataset
    for ci, label in enumerate(["Framework", "Model", "Dataset"], 1):
        ws.merge_cells(f"{get_column_letter(ci)}1:{get_column_letter(ci)}2")
        hdr(ws, 1, ci, label)

    col = 4
    for disp, _, unit, _ in GIN_ACC_METRICS:
        end = get_column_letter(col + 3)
        ws.merge_cells(f"{get_column_letter(col)}1:{end}1")
        hdr(ws, 1, col, disp)
        subhdr(ws, 2, col,   "20 ep (Auto)",   fill=SUBHDR_FILL)
        subhdr(ws, 2, col+1, "300 ep",          fill=EPOCH300_FILL)
        subhdr(ws, 2, col+2, "Î” (300 vs 20)",   fill=SUBHDR_FILL)
        subhdr(ws, 2, col+3, "Ã— gain",          fill=SUBHDR_FILL)
        col += 4

    row = 3
    for mode in COMPILE_MODES:
        # section divider
        last_col = get_column_letter(3 + len(GIN_ACC_METRICS) * 4)
        ws.merge_cells(f"A{row}:{last_col}{row}")
        cell = ws.cell(row, 1, f"â”€â”€ {MODE_SHORT[mode]} â”€â”€")
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = SUBHDR2_FILL; cell.alignment = CENTER; cell.border = BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        for fw, model, ds in all_gin_keys:
            key  = (fw, model, ds)
            ae   = auto_exps.get(key,  {}).get(mode, {})
            g3e  = gin300_exps.get(key, {}).get(mode, {})
            fill = ALT_ROW_FILL if row % 2 == 0 else NO_FILL

            sc(ws.cell(row, 1), fw.upper(),    fill=fill)
            sc(ws.cell(row, 2), model.upper(), fill=fill)
            sc(ws.cell(row, 3), ds,            fill=fill)

            col = 4
            for _, key_j, _, lib in GIN_ACC_METRICS:
                a_err  = ae.get("_error")  if ae  else None
                g3_err = g3e.get("_error") if g3e else None
                av  = ae.get(key_j)  if ae  and not a_err  else None
                g3v = g3e.get(key_j) if g3e and not g3_err else None

                il_a  = is_link_metric(ae,  key_j)
                il_g3 = is_link_metric(g3e, key_j)

                data_cell(ws, row, col,   a_err  or av,  fill=fill, err=bool(a_err),  neutral=not ae,  is_link=il_a)
                data_cell(ws, row, col+1, g3_err or g3v, fill=fill, err=bool(g3_err), neutral=not g3e, is_link=il_g3)
                delta_cell(ws, row, col+2, av, g3v, lib, row_fill=fill)   # Î” 300ep vs 20ep

                gain_cell = ws.cell(row, col+3)
                if isinstance(av,(int,float)) and isinstance(g3v,(int,float)) and av != 0 and g3v != 0 and lib is not None:
                    ratio = (av / g3v) if lib else (g3v / av)   # lower_is_better: 20ep/300ep; higher: 300ep/20ep
                    gain_cell.value     = round(ratio, 3)
                    gain_cell.font      = BODY_FONT
                    gain_cell.fill      = GREEN_FILL if ratio > 1 else (RED_FILL if ratio < 1 else fill)
                    gain_cell.alignment = RIGHT
                else:
                    gain_cell.value = "â€”"; gain_cell.fill = GREY_FILL
                    gain_cell.font = BODY_FONT; gain_cell.alignment = CENTER
                gain_cell.border = BORDER
                col += 4
            row += 1

        # Median row for this mode
        fill = SUMM_FILL
        sc(ws.cell(row, 1), f"MEDIAN â€“ {MODE_SHORT[mode]}", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 2), "", font=SUMM_FONT, fill=fill)
        sc(ws.cell(row, 3), "", font=SUMM_FONT, fill=fill)
        col = 4
        for _, key_j, _, lib in GIN_ACC_METRICS:
            am  = safe_median([get_val(auto_exps,   k, mode, key_j) for k in all_gin_keys])
            g3m = safe_median([get_val(gin300_exps, k, mode, key_j) for k in all_gin_keys])
            data_cell(ws, row, col,   am,  fill=fill)
            data_cell(ws, row, col+1, g3m, fill=fill)
            delta_cell(ws, row, col+2, am, g3m, lib, row_fill=fill)
            ws.cell(row, col+3).value = ""; ws.cell(row, col+3).fill = fill; ws.cell(row, col+3).border = BORDER
            col += 4
        row += 2

    # Legend
    row += 1
    for fill, label in [
        (GREEN_FILL,  "Ã— gain > 1.0 â€” 300-epoch version is better (accuracy higher / latency lower)"),
        (RED_FILL,    "Ã— gain < 1.0 â€” 20-epoch version is better (unusual; may indicate overfit or instability)"),
        (GREY_FILL,   "â€” only one condition has data for this metric"),
    ]:
        c = ws.cell(row, 1, label)
        c.font = BODY_FONT; c.fill = fill; c.alignment = LEFT; c.border = BORDER
        ws.merge_cells(f"A{row}:{get_column_letter(3 + len(GIN_ACC_METRICS)*4)}{row}")
        row += 1

    cw(ws, {1:10, 2:12, 3:18})
    col = 4
    for _ in GIN_ACC_METRICS:
        cw(ws, {col:13, col+1:13, col+2:11, col+3:9}); col += 4



# MAIN
def main():
    parser = argparse.ArgumentParser(
        description="GNN Benchmark Results Analyzer",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--false",    "-f",  required=False, metavar="DIR", default="results_false",
                        help="Root directory with results_false experiment folders")
    parser.add_argument("--true",     "-t",  required=False, metavar="DIR", default="results_true",
                        help="Root directory with results_true experiment folders")
    parser.add_argument("--auto",     "-a",  required=False, metavar="DIR", default="results_auto",
                        help="Root directory with results_auto experiment folders")
    parser.add_argument("--nofixes",  "-n",  required=False, metavar="DIR", default="results_no_fixes",
                        help="Root directory with results_no_fixes experiment folders (no code workarounds)")
    parser.add_argument("--gin300",   "-g",  required=False, metavar="DIR", default="results_gin_300epochs",
                        help="Root directory with results_gin_300epochs (GIN accuracy run, 300 train epochs)")
    parser.add_argument("--output",   "-o",  default="gnn_benchmark_results.xlsx",
                        metavar="FILE", help="Output .xlsx file path")
    args = parser.parse_args()

    false_root   = Path(getattr(args, "false"))
    true_root    = Path(getattr(args, "true"))
    auto_root    = Path(getattr(args, "auto"))
    nofixes_root = Path(args.nofixes) if args.nofixes else None
    gin300_root  = Path(args.gin300)  if args.gin300  else None
    out_path     = Path(args.output)

    if not false_root.exists():
        parser.error(f"False directory not found: {false_root}")
    if not true_root.exists():
        parser.error(f"True directory not found: {true_root}")
    if not auto_root.exists():
        parser.error(f"Auto directory not found: {auto_root}")
    # nofixes and gin300 are optional â€” silently skip if the default folder is absent
    if nofixes_root and not nofixes_root.exists():
        print(f"Note: --nofixes directory not found ({nofixes_root}) â€” skipping no-fixes sheets.")
        nofixes_root = None
    if gin300_root and not gin300_root.exists():
        print(f"Note: --gin300 directory not found ({gin300_root}) â€” skipping GIN accuracy sheet.")
        gin300_root = None

    print("Loading false experiments...")
    false_exps = load_experiments(false_root)
    print(f"  {len(false_exps)} configurations")

    print("Loading true experiments...")
    true_exps = load_experiments(true_root)
    print(f"  {len(true_exps)} configurations")

    print("Loading auto experiments...")
    auto_exps = load_experiments(auto_root)
    print(f"  {len(auto_exps)} configurations")

    if nofixes_root:
        print("Loading no-fixes experiments...")
        nofixes_exps = load_experiments(nofixes_root)
        print(f"  {len(nofixes_exps)} configurations")
    else:
        nofixes_exps = {}
        print("No-fixes directory not provided â€” skipping no-fixes sheets.")

    if gin300_root:
        print("Loading GIN 300-epoch experiments...")
        gin300_exps = load_experiments(gin300_root)
        print(f"  {len(gin300_exps)} configurations")
    else:
        gin300_exps = {}
        print("GIN-300 directory not provided â€” skipping GIN accuracy sheet.")

    wb = Workbook()
    wb.remove(wb.active)

    # Helper wrapper for combined dicts across all frameworks where appropriate
    merged_all = auto_exps

    print("Building Research Questions overview...")
    build_rq_overview_sheet(wb, false_exps, true_exps, auto_exps, nofixes_exps)

    print("Building per-mode comparison sheets (Auto baseline vs True vs False)...")
    for mode in COMPILE_MODES:
        name = "Eager (Baseline)" if mode == "eager" else mode.replace("-", " ").title()
        print(f"  {name}")
        build_comparison_sheet(wb, name, false_exps, true_exps, auto_exps, nofixes_exps, mode)

    print("Building mode summary (median pivot) sheets...")
    build_mode_summary_sheet(wb, auto_exps,    "Mode Summary â€“ Auto")
    build_mode_summary_sheet(wb, true_exps,    "Mode Summary â€“ True")
    build_mode_summary_sheet(wb, false_exps,   "Mode Summary â€“ False")
    if nofixes_exps:
        build_mode_summary_sheet(wb, nofixes_exps, "Mode Summary â€“ No-Fixes")

    print("Building per-framework summary sheets...")
    build_per_framework_summary(wb, false_exps, true_exps, auto_exps, "dgl", "DGL Summary", nofixes_exps)
    build_per_framework_summary(wb, false_exps, true_exps, auto_exps, "pyg", "PyG Summary", nofixes_exps)

    print("Building DGL vs PyG comparison sheets...")
    build_fw_eager_sheet(wb, merged_all)
    build_fw_all_modes_sheet(wb, merged_all)

    print("Building DGL Eager vs PyG Compiled sheet...")
    build_dgl_eager_vs_pyg_compiled_sheet(wb, merged_all)

    print("Building Compile Modes vs Eager (within-framework) sheet...")
    build_compile_vs_eager_sheet(wb, merged_all)

    print("Building Memory Analysis sheets (5 separate sheets)...")
    build_memory_dgl_vs_pyg_sheet(wb, auto_exps)
    build_memory_per_mode_vs_eager_sheet(wb, auto_exps)
    build_memory_pyg_sheet(wb, auto_exps)
    build_memory_dgl_sheet(wb, auto_exps)
    build_memory_best_sheet(wb, auto_exps)

    if nofixes_exps:
        print("Building Auto vs No-Fixes comparison sheet...")
        build_auto_vs_nofixes_sheet(wb, auto_exps, nofixes_exps)

    if gin300_exps:
        print("Building GIN 300 vs 20 epochs accuracy sheet...")
        build_gin_accuracy_sheet(wb, auto_exps, gin300_exps)

    print("Building error summary...")
    build_error_summary_sheet(wb, false_exps, true_exps, auto_exps, nofixes_exps)

    wb.save(out_path)
    print(f"\nâœ“ Saved: {out_path}")
    print("  Sheets:")
    for s in wb.sheetnames:
        print(f"    â€¢ {s}")


if __name__ == "__main__":
    main()
